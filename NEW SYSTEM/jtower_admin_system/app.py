import os, io, hashlib, functools
from datetime import datetime, timedelta
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, send_file, jsonify)
import mysql.connector
from mysql.connector import Error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── App Setup ───────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(__file__)
app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = 'jtower-hoku-secret-2026-xK9#mP'
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
ALLOWED = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# ─── Database Connection ──────────────────────────────────────────────────────
def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',       # Update with your MySQL username
            password='',       # Update with your MySQL password
            database='jtower'
        )
        return conn
    except Error as e:
        print(f"Error connecting to MySQL: {e}")
        return None

# ─── Helpers ─────────────────────────────────────────────────────────────────
def allowed_file(fn): return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def login_required(f):
    @functools.wraps(f)
    def decorated(*a, **kw):
        if 'user_id' not in session:
            flash('Please log in to continue.', 'warning')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return decorated

def next_ref(prefix, table, col):
    conn = get_db_connection()
    if not conn:
        return f"{prefix}-{datetime.now().year}-001"
    cursor = conn.cursor(dictionary=True)
    cursor.execute(f"SELECT {col} FROM {table} ORDER BY id DESC LIMIT 1")
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if row:
        try:
            n = int(row[col].split('-')[-1]) + 1
        except Exception:
            n = 1
    else:
        n = 1
    return f"{prefix}-{datetime.now().year}-{n:03d}"

# ─── Auth ─────────────────────────────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()
            if user and user['password'] == hash_password(password):
                session['user_id']   = user['id']
                session['username']  = user['username']
                session['full_name'] = user['full_name']
                session['role']      = user['role']
                return redirect(url_for('dashboard'))
            error = 'Invalid username or password.'
        else:
            error = 'Database connection failed.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# ─── Dashboard ───────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    period = request.args.get('period', 'monthly').lower()
    if period not in ['weekly', 'monthly', 'yearly']:
        period = 'monthly'

    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger')
        return render_template('dashboard.html', summary={}, chart_data={'labels':[],'datasets':{}}, period=period)
    cursor = conn.cursor(dictionary=True)

    report_types = {
        'incident': 'incident_reports',
        'transmittal': 'transmittal_reports',
        'progress': 'progress_reports',
        'explanation': 'explanation_reports',
        'unit_turnover': 'unit_turnover'
    }

    # 1. Summary Counts (Total Saved Records)
    summary = {}
    for key, table in report_types.items():
        cursor.execute(f"SELECT COUNT(id) as cnt FROM {table}")
        res = cursor.fetchone()
        summary[key] = res['cnt'] if res else 0

    # 2. Unified Graph Data
    # Generate timeline labels
    today = datetime.today()
    labels = []
    keys = [] # used for DB mapping

    if period == 'yearly':
        sql_fmt = "%Y"
        for i in range(4, -1, -1):
            y = today.year - i
            labels.append(str(y))
            keys.append(str(y))
    elif period == 'weekly':
        sql_fmt = "%Y-%u"
        start_of_week = today - timedelta(days=today.weekday())
        for i in range(11, -1, -1):
            dt = start_of_week - timedelta(weeks=i)
            labels.append(dt.strftime('Wk %W, %Y'))
            keys.append(dt.strftime('%Y-%W'))
    else: # monthly
        sql_fmt = "%Y-%m"
        for i in range(11, -1, -1):
            y, m = today.year, today.month - i
            while m <= 0: m += 12; y -= 1
            dt = datetime(y, m, 1)
            labels.append(dt.strftime('%b %Y'))
            keys.append(dt.strftime('%Y-%m'))

    datasets = {}
    for key, table in report_types.items():
        date_col = 'to_date' if key == 'unit_turnover' else 'date'
        query = f"""
            SELECT DATE_FORMAT({date_col}, '{sql_fmt}') as k, COUNT(id) AS c
            FROM {table}
            WHERE {date_col} IS NOT NULL
            GROUP BY k
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        data_map = { r['k']: r['c'] for r in rows }
        # Align with generated keys
        datasets[key] = [data_map.get(k, 0) for k in keys]

    chart_data = { 'labels': labels, 'datasets': datasets }
    cursor.close()
    conn.close()
    return render_template('dashboard.html', summary=summary, chart_data=chart_data, period=period)

# ─── Incident Reports ─────────────────────────────────────────────────────────
@app.route('/incidents')
@login_required
def incidents():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM incident_reports WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (ref_number LIKE %s OR incident_title LIKE %s OR prepared_by LIKE %s OR location LIKE %s)"
        params += [like, like, like, like]
    if month:
        query += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    if not conn:
        return render_template('incidents/list.html', rows=[], search=search, month=month,
                               report_months=[], sort=sort, next_ref='IR-2026-001',
                               today=datetime.now().strftime('%Y-%m-%d'))
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE_FORMAT(date, '%%Y-%%m') as month FROM incident_reports WHERE date IS NOT NULL ORDER BY month DESC")
    report_months = []
    for r in cursor.fetchall():
        try:
            dt_obj = datetime.strptime(r['month'], '%Y-%m')
            report_months.append({'value': r['month'], 'display': dt_obj.strftime('%B %Y')})
        except (ValueError, TypeError):
            continue
    next_ir_ref = next_ref('IR', 'incident_reports', 'ref_number')
    cursor.close(); conn.close()
    return render_template('incidents/list.html', rows=rows, search=search, month=month,
                           report_months=report_months, sort=sort, next_ref=next_ir_ref,
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/incidents/new', methods=['GET', 'POST'])
@login_required
def incident_new():
    if request.method == 'POST':
        image_path = None
        f = request.files.get('image')
        if f and f.filename and allowed_file(f.filename):
            ext   = f.filename.rsplit('.', 1)[1].lower()
            fname = f"IR_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
            image_path = fname
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            inserted = False
            attempts = 0
            while not inserted and attempts < 3:
                try:
                    ref_num = next_ref('IR', 'incident_reports', 'ref_number')
                    cursor.execute("""
                        INSERT INTO incident_reports
                          (ref_number,date,location,incident_title,department,image_path,
                           prepared_by,noted_by,received_by,description,status)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        ref_num, request.form['date'],
                        request.form['location'], request.form['incident_title'],
                        request.form['department'], image_path,
                        request.form['prepared_by'],
                        request.form.get('noted_by', 'Michael P. Resilosa'),
                        request.form.get('received_by', ''),
                        request.form.get('description', ''),
                        request.form.get('status', 'Open'),
                    ))
                    conn.commit()
                    inserted = True
                except mysql.connector.Error as e:
                    if e.errno == 1062: # Duplicate entry
                        attempts += 1
                        conn.rollback()
                    else:
                        cursor.close(); conn.close()
                        raise e
            
            cursor.close(); conn.close()
            if not inserted:
                flash('Error creating report: Could not generate unique reference number.', 'danger')
                return redirect(url_for('incidents'))

        flash('Incident report created successfully.', 'success')
        return redirect(url_for('incidents'))
    ref = next_ref('IR', 'incident_reports', 'ref_number')
    return render_template('incidents/form.html', record=None, ref=ref,
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/incidents/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def incident_edit(rid):
    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger')
        return redirect(url_for('incidents'))
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM incident_reports WHERE id=%s", (rid,))
    rec = cursor.fetchone()
    if not rec:
        cursor.close(); conn.close()
        flash('Record not found.', 'danger')
        return redirect(url_for('incidents'))
    if request.method == 'POST':
        image_path = rec['image_path']
        f = request.files.get('image')
        if f and f.filename and allowed_file(f.filename):
            ext   = f.filename.rsplit('.', 1)[1].lower()
            fname = f"IR_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
            image_path = fname
        cursor.execute("""
            UPDATE incident_reports SET
              date=%s, location=%s, incident_title=%s, department=%s,
              image_path=%s, prepared_by=%s, noted_by=%s, received_by=%s,
              description=%s, status=%s, updated_at=NOW()
            WHERE id=%s
        """, (
            request.form['date'], request.form['location'],
            request.form['incident_title'], request.form['department'],
            image_path, request.form['prepared_by'],
            request.form.get('noted_by', ''), request.form.get('received_by', ''),
            request.form.get('description', ''), request.form.get('status', 'Open'), rid
        ))
        conn.commit(); cursor.close(); conn.close()
        flash('Incident report updated.', 'success')
        return redirect(url_for('incidents'))
    cursor.close(); conn.close()
    return render_template('incidents/form.html', record=rec, ref=rec['ref_number'],
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/incidents/<int:rid>/delete', methods=['POST'])
@login_required
def incident_delete(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM incident_reports WHERE id=%s", (rid,))
        conn.commit(); cursor.close(); conn.close()
    flash('Record deleted.', 'info')
    return redirect(url_for('incidents'))

@app.route('/incidents/export')
@login_required
def incident_export():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM incident_reports WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (ref_number LIKE %s OR incident_title LIKE %s OR prepared_by LIKE %s OR location LIKE %s)"
        params += [like, like, like, like]
    if month:
        query += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    rows = []
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        cursor.close(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Incident Reports"
    navy = PatternFill("solid", fgColor="0F172A")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    title_font = Font(bold=True, name="Calibri", size=14, color="0F172A")
    thin = Border(left=Side(style='thin',color='D1D5DB'), right=Side(style='thin',color='D1D5DB'),
                  top=Side(style='thin',color='D1D5DB'),  bottom=Side(style='thin',color='D1D5DB'))
    ws.merge_cells('A1:I1'); ws['A1'] = "J Tower Residences – Incident Reports"
    ws['A1'].font = title_font; ws['A1'].fill = PatternFill("solid", fgColor="FFF6E0")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 28
    ws.merge_cells('A2:I2'); ws['A2'] = f"Exported: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
    ws['A2'].font = Font(italic=True, color="888888", size=9); ws['A2'].alignment = Alignment(horizontal='center')
    headers = ['Ref Number','Date','Location','Incident Title','Department','Image','Prepared By','Status','Created At']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hdr_font; cell.fill = navy
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin
    ws.row_dimensions[3].height = 22
    alt = PatternFill("solid", fgColor="F8F9FA")
    for i, row in enumerate(rows, 4):
        fill = alt if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        data = [row['ref_number'], str(row['date']) if row['date'] else '', row['location'],
                row['incident_title'], row['department'],
                '(attached)' if row['image_path'] else 'None',
                row['prepared_by'], row['status'], str(row['created_at']) if row['created_at'] else '']
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill; cell.border = thin; cell.alignment = Alignment(vertical='center', wrap_text=True)
    for i, w in enumerate([16,13,22,30,16,12,20,14,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"Incident_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ─── Transmittal Reports ──────────────────────────────────────────────────────
@app.route('/transmittal')
@login_required
def transmittal():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM transmittal_reports WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (ref_number LIKE %s OR subject LIKE %s OR to_party LIKE %s OR from_party LIKE %s OR prepared_by LIKE %s)"
        params += [like, like, like, like, like]
    if month:
        query += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    if not conn:
        return render_template('transmittal.html', rows=[], search=search, month=month,
                               sort=sort, report_months=[], next_ref='TR-2026-001',
                               today=datetime.now().strftime('%Y-%m-%d'))
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE_FORMAT(date, '%%Y-%%m') as month FROM transmittal_reports WHERE date IS NOT NULL ORDER BY month DESC")
    report_months = []
    for r in cursor.fetchall():
        try:
            dt_obj = datetime.strptime(r['month'], '%Y-%m')
            report_months.append({'value': r['month'], 'display': dt_obj.strftime('%B %Y')})
        except (ValueError, TypeError):
            continue
    next_tr_ref = next_ref('TR', 'transmittal_reports', 'ref_number')
    cursor.close(); conn.close()
    return render_template('transmittal.html', rows=rows, search=search, month=month,
                           sort=sort, report_months=report_months, next_ref=next_tr_ref,
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/transmittal/new', methods=['GET', 'POST'])
@login_required
def transmittal_new():
    if request.method == 'POST':
        image_path = None
        f = request.files.get('image')
        if f and f.filename and allowed_file(f.filename):
            ext   = f.filename.rsplit('.', 1)[1].lower()
            fname = f"TR_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
            image_path = fname
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO transmittal_reports
                  (ref_number,date,to_party,from_party,subject,document_count,prepared_by,received_by,remarks,status,image_path)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                request.form['ref_number'], request.form['date'],
                request.form['to_party'], request.form['from_party'],
                request.form['subject'], request.form.get('document_count', 1),
                request.form['prepared_by'], request.form.get('received_by', ''),
                request.form.get('remarks', ''), request.form.get('status', 'Pending'), image_path
            ))
            conn.commit(); cursor.close(); conn.close()
        flash('Transmittal report created.', 'success')
        return redirect(url_for('transmittal'))
    ref = next_ref('TR', 'transmittal_reports', 'ref_number')
    return render_template('transmittal_form.html', record=None, ref=ref,
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/transmittal/<int:rid>/edit', methods=['POST'])
@login_required
def transmittal_edit(rid):
    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger'); return redirect(url_for('transmittal'))
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM transmittal_reports WHERE id=%s", (rid,))
    rec = cursor.fetchone()
    if not rec:
        cursor.close(); conn.close(); flash('Record not found.', 'danger'); return redirect(url_for('transmittal'))
    image_path = rec.get('image_path')
    f = request.files.get('image')
    if f and f.filename and allowed_file(f.filename):
        ext   = f.filename.rsplit('.', 1)[1].lower()
        fname = f"TR_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        f.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
        image_path = fname
    cursor.execute("""
        UPDATE transmittal_reports SET
          date=%s, to_party=%s, from_party=%s, subject=%s,
          document_count=%s, prepared_by=%s, received_by=%s,
          remarks=%s, status=%s, image_path=%s, updated_at=NOW()
        WHERE id=%s
    """, (
        request.form['date'], request.form.get('to_party', ''),
        request.form['from_party'], request.form['subject'],
        request.form.get('document_count', 1), request.form['prepared_by'],
        request.form.get('received_by', ''), request.form.get('remarks', ''),
        request.form.get('status', 'Pending'), image_path, rid
    ))
    conn.commit(); cursor.close(); conn.close()
    flash('Transmittal report updated.', 'success')
    return redirect(url_for('transmittal'))

@app.route('/transmittal/<int:rid>/delete', methods=['POST'])
@login_required
def transmittal_delete(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM transmittal_reports WHERE id=%s", (rid,))
        conn.commit(); cursor.close(); conn.close()
    flash('Transmittal report deleted.', 'info')
    return redirect(url_for('transmittal'))

# ─── Progress Reports ─────────────────────────────────────────────────────────
@app.route('/progress')
@login_required
def progress():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM progress_reports WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (ref_number LIKE %s OR employee_name LIKE %s OR reporting_period LIKE %s)"
        params += [like, like, like]
    if month:
        query += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    if not conn:
        return render_template('progress.html', rows=[], search=search, month=month,
                               sort=sort, report_months=[])
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE_FORMAT(date, '%%Y-%%m') as month FROM progress_reports WHERE date IS NOT NULL ORDER BY month DESC")
    report_months = []
    for r in cursor.fetchall():
        try:
            dt_obj = datetime.strptime(r['month'], '%Y-%m')
            report_months.append({'value': r['month'], 'display': dt_obj.strftime('%B %Y')})
        except (ValueError, TypeError):
            continue
    next_pr_ref = next_ref('PR', 'progress_reports', 'ref_number')
    cursor.close(); conn.close()
    return render_template('progress.html', rows=rows, search=search, month=month,
                           sort=sort, report_months=report_months, next_ref=next_pr_ref)

@app.route('/progress/new', methods=['GET', 'POST'])
@login_required
def progress_new():
    if request.method == 'POST':
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            
            # Robust generation of Ref Number with retry on collision
            inserted = False
            attempts = 0
            while not inserted and attempts < 3:
                try:
                    # Generate fresh ref number
                    ref_num = next_ref('PR', 'progress_reports', 'ref_number')
                    cursor.execute("""
                        INSERT INTO progress_reports
                          (ref_number, date, department, employee_name, employee_position, reporting_period,
                           work_accomplishments, tasks_completed, issues_concerns, actions_taken, 
                           planned_tasks, remarks, reviewed_by, reviewer_position)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        ref_num, request.form['date'],
                        request.form['department'], request.form.get('employee_name',''),
                        request.form.get('employee_position',''), request.form.get('reporting_period',''),
                        request.form.get('work_accomplishments',''), request.form.get('tasks_completed',''),
                        request.form.get('issues_concerns',''), request.form.get('actions_taken',''),
                        request.form.get('planned_tasks',''), request.form.get('remarks',''),
                        request.form.get('reviewed_by',''), request.form.get('reviewer_position','')
                    ))
                    conn.commit()
                    inserted = True
                except mysql.connector.Error as e:
                    if e.errno == 1062: # Duplicate entry code
                        attempts += 1
                        conn.rollback() # Rollback to retry
                    else:
                        cursor.close(); conn.close()
                        raise e
            
            cursor.close(); conn.close()
            if not inserted:
                flash('Error creating report: Could not generate unique reference number.', 'danger')
                return redirect(url_for('progress'))

        flash('Progress report created.', 'success')
        return redirect(url_for('progress'))
    ref = next_ref('PR', 'progress_reports', 'ref_number')
    return render_template('progress_form.html', record=None, ref=ref,
                           today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/progress/<int:rid>/edit', methods=['POST'])
@login_required
def progress_edit(rid):
    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger'); return redirect(url_for('progress'))
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM progress_reports WHERE id=%s", (rid,))
    rec = cursor.fetchone()
    if not rec:
        cursor.close(); conn.close(); flash('Record not found.', 'danger'); return redirect(url_for('progress'))
    cursor.execute("""
        UPDATE progress_reports SET
          date=%s, department=%s, employee_name=%s, employee_position=%s,
          reporting_period=%s, work_accomplishments=%s, tasks_completed=%s,
          issues_concerns=%s, actions_taken=%s, planned_tasks=%s, remarks=%s,
          reviewed_by=%s, reviewer_position=%s, updated_at=NOW()
        WHERE id=%s
    """, (
        request.form['date'], request.form['department'], request.form.get('employee_name',''),
        request.form.get('employee_position',''), request.form.get('reporting_period',''),
        request.form.get('work_accomplishments',''), request.form.get('tasks_completed',''),
        request.form.get('issues_concerns',''), request.form.get('actions_taken',''),
        request.form.get('planned_tasks',''), request.form.get('remarks',''),
        request.form.get('reviewed_by',''), request.form.get('reviewer_position',''), rid
    ))
    conn.commit(); cursor.close(); conn.close()
    flash('Progress report updated.', 'success')
    return redirect(url_for('progress'))

@app.route('/progress/<int:rid>/delete', methods=['POST'])
@login_required
def progress_delete(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM progress_reports WHERE id=%s", (rid,))
        conn.commit(); cursor.close(); conn.close()
    flash('Progress report deleted.', 'info')
    return redirect(url_for('progress'))

# ─── Explanation Reports ──────────────────────────────────────────────────────
@app.route('/explanation')
@login_required
def explanation():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM explanation_reports WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (ref_number LIKE %s OR location LIKE %s OR prepared_by LIKE %s)"
        params += [like, like, like]
    if month:
        query += " AND DATE_FORMAT(date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    if not conn:
        return render_template('explanation.html', rows=[], search=search, month=month,
                               sort=sort, report_months=[])
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE_FORMAT(date, '%%Y-%%m') as month FROM explanation_reports WHERE date IS NOT NULL ORDER BY month DESC")
    report_months = []
    for r in cursor.fetchall():
        try:
            dt_obj = datetime.strptime(r['month'], '%Y-%m')
            report_months.append({'value': r['month'], 'display': dt_obj.strftime('%B %Y')})
        except (ValueError, TypeError):
            continue
    next_ex_ref = next_ref('ER', 'explanation_reports', 'ref_number')
    cursor.close(); conn.close()
    return render_template('explanation.html', rows=rows, search=search, month=month,
                           sort=sort, report_months=report_months, next_ref=next_ex_ref)

@app.route('/explanation/new', methods=['POST'])
@login_required
def explanation_new():
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO explanation_reports
              (ref_number,date,location,department,description,prepared_by,received_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form['ref_number'], request.form['date'],
            request.form['location'], request.form['department'],
            request.form.get('description', ''),
            request.form['prepared_by'],
            request.form.get('received_by', ''),
        ))
        conn.commit(); cursor.close(); conn.close()
    flash('Explanation report created.', 'success')
    return redirect(url_for('explanation'))

@app.route('/explanation/<int:rid>/edit', methods=['POST'])
@login_required
def explanation_edit(rid):
    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger'); return redirect(url_for('explanation'))
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE explanation_reports SET
          date=%s, location=%s, department=%s, description=%s,
          prepared_by=%s, received_by=%s, updated_at=NOW()
        WHERE id=%s
    """, (
        request.form['date'], request.form['location'], request.form['department'],
        request.form.get('description', ''),
        request.form['prepared_by'], request.form.get('received_by', ''), rid
    ))
    conn.commit(); cursor.close(); conn.close()
    flash('Explanation report updated.', 'success')
    return redirect(url_for('explanation'))

@app.route('/explanation/<int:rid>/delete', methods=['POST'])
@login_required
def explanation_delete(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM explanation_reports WHERE id=%s", (rid,))
        conn.commit(); cursor.close(); conn.close()
    flash('Explanation report deleted.', 'info')
    return redirect(url_for('explanation'))

# ─── Unit Turn Over ───────────────────────────────────────────────────────────
@app.route('/unit-turnover')
@login_required
def unit_turnover():
    search = request.args.get('search', '').strip()
    month  = request.args.get('month', '').strip()
    sort   = request.args.get('sort', 'latest')
    query  = "SELECT * FROM unit_turnover WHERE 1=1"
    params = []
    if search:
        like = f"%{search}%"
        query += " AND (unit_number LIKE %s OR unit_owner_name LIKE %s OR control_number LIKE %s)"
        params += [like, like, like]
    if month:
        query += " AND DATE_FORMAT(to_date, '%%Y-%%m') = %s"
        params.append(month)
    query += " ORDER BY created_at " + ("ASC" if sort == 'oldest' else "DESC")
    conn = get_db_connection()
    if not conn:
        flash('Database connection failed.', 'danger')
        return render_template('unit_turnover.html', rows=[], search=search, month=month,
                               sort=sort, report_months=[])
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.execute("SELECT DISTINCT DATE_FORMAT(to_date, '%%Y-%%m') as month FROM unit_turnover WHERE to_date IS NOT NULL ORDER BY month DESC")
    report_months = []
    for r in cursor.fetchall():
        try:
            dt_obj = datetime.strptime(r['month'], '%Y-%m')
            report_months.append({'value': r['month'], 'display': dt_obj.strftime('%B %Y')})
        except (ValueError, TypeError):
            continue

    # Custom generation for JTRCC-CTO-00001 format
    cursor.execute("SELECT control_number FROM unit_turnover WHERE control_number LIKE 'JTRCC-CTO-%' ORDER BY id DESC LIMIT 1")
    last_rec = cursor.fetchone()
    uto_seq = 1
    if last_rec and last_rec['control_number']:
        try:
            parts = last_rec['control_number'].rsplit('-', 1)
            if len(parts) == 2 and parts[1].isdigit():
                uto_seq = int(parts[1]) + 1
        except Exception:
            pass
    next_uto_ref = f"JTRCC-CTO-{uto_seq:05d}"

    cursor.close(); conn.close()
    return render_template('unit_turnover.html', rows=rows, search=search, month=month,
                           sort=sort, report_months=report_months, next_ref=next_uto_ref)

@app.route('/unit-turnover/new', methods=['POST'])
@login_required
def unit_turnover_new():
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO unit_turnover
              (control_number, unit_number, unit_owner_name, to_date, unit_type, tower_floor)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            request.form['control_number'], request.form['unit_number'],
            request.form['unit_owner_name'], request.form['to_date'],
            request.form['unit_type'], request.form.get('tower_floor', '')
        ))
        conn.commit(); cursor.close(); conn.close()
    flash('Unit turnover record created.', 'success')
    return redirect(url_for('unit_turnover'))

@app.route('/unit-turnover/<int:rid>/edit', methods=['POST'])
@login_required
def unit_turnover_edit(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE unit_turnover SET
              control_number=%s, unit_number=%s, unit_owner_name=%s,
              to_date=%s, unit_type=%s, tower_floor=%s, updated_at=NOW()
            WHERE id=%s
        """, (
            request.form['control_number'], request.form['unit_number'],
            request.form['unit_owner_name'], request.form['to_date'],
            request.form['unit_type'], request.form.get('tower_floor', ''), rid
        ))
        conn.commit(); cursor.close(); conn.close()
    flash('Unit turnover record updated.', 'success')
    return redirect(url_for('unit_turnover'))

@app.route('/unit-turnover/<int:rid>/delete', methods=['POST'])
@login_required
def unit_turnover_delete(rid):
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM unit_turnover WHERE id=%s", (rid,))
        conn.commit(); cursor.close(); conn.close()
    flash('Unit turnover record deleted.', 'info')
    return redirect(url_for('unit_turnover'))

# ─── Settings ─────────────────────────────────────────────────────────────────
@app.route('/settings/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not all([current_password, new_password, confirm_password]):
            flash('All fields are required.', 'danger')
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('change_password'))

        conn = get_db_connection()
        if not conn:
            flash('Database connection failed.', 'danger')
            return redirect(url_for('change_password'))

        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()

        if user and user['password'] == hash_password(current_password):
            new_hashed_password = hash_password(new_password)
            cursor.execute("UPDATE users SET password = %s WHERE id = %s", (new_hashed_password, session['user_id']))
            conn.commit()
            flash('Your password has been updated successfully.', 'success')
        else:
            flash('Incorrect current password.', 'danger')

        cursor.close(); conn.close()
        return redirect(url_for('change_password'))

    return render_template('settings.html')

# ─── Run ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("J Tower Admin System running -> http://127.0.0.1:5000")
    print("   Default login  admin / admin123")
    app.run(debug=True, port=5000)
